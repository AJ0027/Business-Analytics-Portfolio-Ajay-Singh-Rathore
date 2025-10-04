from openpyxl import load_workbook
from datetime import datetime

# Load the workbook and sheet
file_path = "sagatave_eksamenam.xlsx"
wb = load_workbook(file_path)
ws = wb["Lapa_0"]

# Extract actual headers from row 3
headers = [cell.value for cell in ws[3]]

# Extract data from row 4 onward
data = []
for row in ws.iter_rows(min_row=4, values_only=True):
    if any(cell is not None for cell in row):
        data.append(row)

# Map data to rows of dictionaries for easy access
records = [dict(zip(headers, row)) for row in data]

# ------------------- Task 1 -------------------
task1 = sum(
    1 for row in records
    if isinstance(row.get("Adrese"), str) and row["Adrese"].startswith("Ain")
    and isinstance(row.get("Skaits"), (int, float)) and row["Skaits"] < 40
)
print("Task 1 Answer:", task1)

# ------------------- Task 2 -------------------
task2 = 0
for row in records:
    priorit = str(row.get("Prioritāte", "")).strip().lower()
    date_str = row.get("Piegādes datums")
    if isinstance(date_str, datetime) and priorit == "high":
        if date_str.year == 2015:
            task2 += 1
print("Task 2 Answer:", task2)

# ------------------- Task 3 -------------------
task3 = sum(
    1 for row in records
    if row.get("Adrese") == "Adulienas iela"
    and row.get("Pilsēta") in ["Valmiera", "Saulkrasti"]
)
print("Task 3 Answer:", task3)

# ------------------- Task 4 -------------------
laserjet_prices = [
    row["Cena"]
    for row in records
    if isinstance(row.get("Cena"), (int, float))
    and isinstance(row.get("Produkts"), str)
    and "LaserJet" in row["Produkts"]
]
task4 = int(sum(laserjet_prices) / len(laserjet_prices)) if laserjet_prices else 0
print("Task 4 Answer:", task4)

# ------------------- Task 5 -------------------
task5_total = 0
for row in records:
    if row.get("Klients") == "Korporatīvais" and isinstance(row.get("Skaits"), (int, float)):
        skaits = row["Skaits"]
        if 40 <= skaits <= 50:
            cena = row.get("Cena", 0) or 0
            pieg_cena = row.get("Piegādes cena", 0) or 0
            kopā = cena * skaits + pieg_cena
            task5_total += kopā
print("Task 5 Answer:", int(task5_total))
